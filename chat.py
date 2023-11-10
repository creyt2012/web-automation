import openai
import pandas as pd
import os
import requests
import random
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm, trange 
import sys
import art

# Function to check if API response is successful
def is_api_response_successful(response):
    if response.status_code != 200:
        print(f"\033[91mLỗi: {response.status_code}, {response.text}\033[0m")
        return False
    return True

# Function to generate text using OpenAI API
def generate_text(prompt, api_key, creativity=0.5):
    endpoint = "https://api.openai.com/v1/engines/text-davinci-003/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    data = {
        "prompt": prompt,
        "max_tokens": 500,
        "temperature": creativity
    }

    response = requests.post(endpoint, headers=headers, json=data)
    if not is_api_response_successful(response):
        return None

    result = response.json()
    return result["choices"][0]["text"]

# Function to append DataFrame to Excel file
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and creating a new workbook
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
    else:
        # Excel file exists
        book = load_workbook(filename)

        # Truncate sheet
        if truncate_sheet and sheet_name in book.sheetnames:
            idx = book.sheetnames.index(sheet_name)
            book.remove(book.worksheets[idx])
            book.create_sheet(sheet_name, idx)

        
        startrow = 0 if startrow is None else startrow
        sheet = book[sheet_name]
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)

        # Save the workbook
        book.save(filename)

def waifu_ascii():
    return """
  
WWWMWWWMWWWWWWWWWWWWWWWWWWWWWMMWWWWWWWMMMWWMMMMWMMMMMMMWMMMMMMMMWWMMMW
WWWWWWWWWWWWWWWWWWWWWWWWWWWMMMMMMMWWWMMMMMMMMMMMMMMMMMMMMMMMWWWWWWWMMM
NNNWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWMMMMMMMMMMMMMWWWWWWWWWWMMM
WWWNNWNNWWWWWWWNNNNWWWWWWNXXK000000000OO000KKXXNWWMWWWMWWWWWWWWWWWWMMM
WWNNNNNWWWWWWWNWWNNNXXK0OkkkkkkkkkkkkkOOOOOOOO000KXNWWWWWWWWWWNNNWWMMM
WWWWNNNWWWWWWWNNNXK0OkkxxxkkkkkkkkkkkkkkkkOOO0000000KXNWWWWWWWWNWWMMMM
WWWWWWWWWWWWNNXK00OOkkxxkkkkkkkkkkkkkkkkxxkkOOO0000000KXNWWWWWWWWWMMMM
WWWWWWNWWWWNXXK00OOkkkkkkkkkkkkkkkkkkkkkkkkxxkOOO0000KXXXNWWWWWWWMMMMM
WWWWWWNWWNNXKK0OOOOOOOOkkkkkOkkkOkkkkkkkkkOkkxxkO0O0KKXNXKXWWWWWMMMMMM
WWWWWWWWNNXK00OOOOOOOOkkkkkkkOkkkkkxxkOOkkOOOkxxkOO0KXXKXXKNWWMMMMMMMM
WWWWWWWNNNX0O00OOO00OkkkkOOkkOOOOkkxdxO0OkkOOO0Oxk0KKXXK0KK0NWWMMMMMMM
WWMWWWNNNNX00K0OO0K0OkOOO00OkOO00OOOxdx0K0OO0K00OxkKKKKK00000NWWMMMMMM
WWWNNXXNNNXKXK0OOXX0kk00KXKOk0000K00Odox0KOkOOOOOxdk0O0000OOOKWWWMMMMM
WWWWNXNNXXKXXK0kOXX0kk000KKOxk0OO000Okdox00kkkOOkkodkkOOOOOOkOXWWWWWWW
WWNWNNNXK00XK0kxOX0OxxOOkO0OddkOkOOOOOkdoxOkkkkkkkdoxkkkkkkkkk0NWWWNNN
WWWNNNX00O0K0OxdOKOkddkOdxOkxddxkkkkkkkxooxkkxxkkkdlokkkkxkkkkkKWWWNNN
WWWNNN0OOO00Oxook0Oxodkkdxkkxdodxxxdxxkxdldxxxxkxxdloxkkxxkkxkx0NNNNNN
WWWNNXOkOOOOkdodkkkxokOddkOxxdldxooooodddllxxddxxxolldxxxxxxxxxkXWNNNW
NWWNNKkkkkkkdlldxdxddO0xoOX0xdook0kdoollolcoxddxddllldxxxxxxxxxd0NNNNN
NNWNN0kxxxxkocckOddddOXOoxXNX0xod0XX0kdlclclxdoddoclloxxxxxxxxxdkXNNNW
WNWNXKkdxxxxollkKOdodONXkxKNNNXOddxkkOkxdocldooddlccclddxxdxddxddONNNN
NNNNNKOddxddoodxxddloONWXO0NWN0xdooo;.',::coddoddlcoxxddddddddxdxOKXXN
WWNNNX0doddoolccloc,;xNWWNXNWNKKXkll;..,;clcododdox00OdodddddddddOXXXX
WWNNWNKxooolll:xKd;''dNWWWWWWWWWWOc::;:cck0dooododkOO0kdoddddddxdxKNNN
WWWNNNN0ololcld0Nxc:cxXWWWWWWWWWWXOxO0KOOXNOoodolx0O0Ooloddxdddddd0NNN
WWWNNNNXkollc:l0NX0O0KNWWWWWWWWWNNXXXXXXNNKxdddolx0Oxc;:ldddxxddxdkXNN
MWWNNNNNKxocc::kXNNXXNNNWWWWWWWWNNNWNXXXNX0Oddddloxo;',;:ldddxddxxx0NN
MMWWNNNWKkxdlc:dXNNNWNXNWWWWWWWWWNNWNNNNNNXOddxoc;;,;'';;codoxxddxdxKN
MMMWWNNWNOxxxo:lKNNWWWWWWWWWWWWWWWWWWWWWWWX0xddoc,,,;,'';:coddxxdxxdOX
MMMWWNNWNKkxxdc:xXWWWWWWWWWWWWWWWWWWWWWWWWN0xddolc,,,;,',;:ldddxxxxxxO
WWWWWWNNNK0Oxoc::dKWWWWWWNXKXNWWWWWWWWWWWNKkdodoxko:;;,,.,::lddddddxxd
WWWWWWNNNX0Okdc;:dKXXNWNWWX00XWWWWWWWWNXKOxxxdook0Odll:;,,;::loddddxkx
NNWWNNNNNX0kxd:;dNWWNXKXXNWWWWWNWNNNXK0OOkkkkdodO0kxxkxoc;,;:ccodoodxk
NNNNNNNNNX0kxo:oKMWWWNK000KXXXX0kkOOOOOOOOkkkdldOOkxkOOkxo:,,:ccloolox
NNNNNNNNNKOkolo0WWMWWWNKK0O000Oo;lk00000OOkOOdoxOkkxkOOOOkkxddoc:ldolo
NNNNNNNNN0kdcoKWMWWWWWWNK0O0KOocoxkOO0000OOOkdoxOOOkxOOOO0KNWNKxc;cddo
NNNNNNNNXkdlcOWWWMWWWWWWNKOkOdlo0K00KKXXXKOOxdok0OOkxO00XWWX0kdc;,,:lx
WWNNNNNN0xlckNWWWMWMMWWMWXkddddOKKKXNNNXXKOkxddk0OOkdOXNWXOxollc;''';c
WWWWNNNNOocdXWWWWWWMWWWWWNkllox0XNNNNNNNNKOkxddkOOOkd0WN0xoloolc::;,,;
MWWWWNNXkcl0NNWWMWWMWNNWWNkcclkKXNWWWWWNNKOkxddkOOOOx0XOdllodkxdool:::
"""
waifu_ascii_result = waifu_ascii()

# Main function
def main():
    
    print("Chào mừng bạn đến với chương trình của tôi!")
    
    print(waifu_ascii_result)
    apikey_file = input("\033[92mNhập đường dẫn đến file txt chứa danh sách API keys:\033[0m ")
    input_file_path = input("\033[92mNhập đường dẫn đến file input XLSX:\033[0m ")  
    output_file_path = input("\033[92mNhập đường dẫn đến file output XLSX:\033[0m ")  
    creativity = float(input("\033[92mNhập mức độ sáng tạo của API (từ 0.0 đến 0.9):\033[0m ")) 
    
    sheet_name = pd.ExcelFile(input_file_path).sheet_names[0]

    with open(apikey_file, 'r') as f:
        api_keys = f.read().splitlines()

    processed_keywords = set()

    if os.path.exists(output_file_path):
        output_data = pd.read_excel(output_file_path, sheet_name=sheet_name, header=None, names=["keyword", "content"])
        processed_keywords.update(output_data["keyword"])

    if not processed_keywords:
        print("\033[91mKhông có dữ liệu để xử lý.\033[0m")  
        return

    api_key_index = 0
    tasks_processed = 0

    data = pd.read_excel(input_file_path, sheet_name=sheet_name)
    data = data.rename(columns={"heading": "keyword"})
    data.to_excel(input_file_path, sheet_name=sheet_name, index=False)

    character_before_heading = input("\033[92mNhập ký tự trước {keyword} (ví dụ form: Viết bài viết chuẩn seo về {keyword}
list keyword này được thêm vào file input:\033[0m ").strip() 

    for _, row in tqdm(data.iterrows(), desc="\033[92mXử lý tác vụ\033[0m", dynamic_ncols=True, total=len(data)):
        keyword = row["keyword"]

        if keyword in processed_keywords:
            print(f"\033[92mKeyword '{keyword}' đã được xử lý trước đó, tiếp tục xử lý...\033[0m") 
        else:
            api_key = api_keys[api_key_index]
            prompt = f"{character_before_heading} {keyword}" if character_before_heading else keyword

            with tqdm(total=100, desc=f"\033[92mĐang xử lý keyword: {keyword}\033[0m", dynamic_ncols=True, postfix=dict, bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]') as pbar:
                content = generate_text(prompt, api_key, creativity)

                if content is not None:
                   
                    output_df = pd.DataFrame([{"keyword": keyword, "content": content}])

                 
                    append_df_to_excel(output_file_path, output_df, sheet_name=sheet_name, startrow=None, index=False, header=False)
                    print(f"\033[92mĐã ghi dữ liệu thành công vào file Excel cho keyword: {keyword}\033[0m")  
                    time.sleep(3)
                    
                    tasks_processed = api_key_index
                    api_key_index = (api_key_index + 1) % len(api_keys)
                    print(f"\033[92mĐã đổi sang API key mới: {api_keys[api_key_index]}\033[0m") 

                    # Cập nhật danh sách processed_keywords
                    processed_keywords.add(keyword)

    # Ghi dữ liệu vào file Excel sau khi đã xử lý tất cả các từ khóa
    output_df = pd.DataFrame(output_data)
    append_df_to_excel(output_file_path, output_df, sheet_name=sheet_name, startrow=None, index=False, header=False)
    print("\033[92mĐã ghi dữ liệu thành công vào file Excel.\033[0m")  

if __name__ == "__main__":
    main()
